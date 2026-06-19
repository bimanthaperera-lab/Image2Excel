from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def rows_to_workbook(columns: list[str], rows: list[list[str]], raw_text: str) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OCR Rows"

    column_count = max((len(row) for row in rows), default=1)
    headers = columns or [f"Column {index}" for index in range(1, column_count + 1)]
    column_count = max(column_count, len(headers), 1)
    sheet.append(headers + [""] * (column_count - len(headers)))
    for row in rows:
        sheet.append((row or [""]) + [""] * (column_count - len(row)))

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    text_sheet = workbook.create_sheet("Raw Text")
    text_sheet["A1"] = raw_text
    text_sheet.column_dimensions["A"].width = 100

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
