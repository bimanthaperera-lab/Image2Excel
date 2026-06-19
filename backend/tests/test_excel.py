from io import BytesIO
import openpyxl
from backend.app.excel import rows_to_workbook

def test_rows_to_workbook_basic():
    columns = ["Name", "Age"]
    rows = [["Alice", "30"], ["Bob", "25"]]
    raw_text = "Alice 30\nBob 25"

    output = rows_to_workbook(columns, rows, raw_text)
    
    assert isinstance(output, BytesIO)
    
    # Check if the workbook is valid
    wb = openpyxl.load_workbook(output)
    
    assert "OCR Rows" in wb.sheetnames
    assert "Raw Text" in wb.sheetnames
    
    sheet = wb["OCR Rows"]
    
    # Check headers
    assert sheet["A1"].value == "Name"
    assert sheet["B1"].value == "Age"
    
    # Check rows
    assert sheet["A2"].value == "Alice"
    assert sheet["B2"].value == "30"
    assert sheet["A3"].value == "Bob"
    assert sheet["B3"].value == "25"
    
    # Check raw text sheet
    text_sheet = wb["Raw Text"]
    assert text_sheet["A1"].value == raw_text

def test_rows_to_workbook_empty():
    output = rows_to_workbook([], [], "")
    wb = openpyxl.load_workbook(output)
    sheet = wb["OCR Rows"]
    
    assert sheet["A1"].value == "Column 1"
